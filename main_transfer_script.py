import pandas as pd
import os
from datetime import datetime
import itertools
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

warehouse_path = r"C:\Users\amira\synthetic_data\warehouse.xlsx"

branch_files = [
    r"C:\Users\amira\synthetic_data\branch_1.xlsx",
    r"C:\Users\amira\synthetic_data\branch_2.xlsx",
    r"C:\Users\amira\synthetic_data\branch_3.xlsx",
    r"C:\Users\amira\synthetic_data\branch_4.xlsx",
    r"C:\Users\amira\synthetic_data\branch_5.xlsx"
]

MIN_WAREHOUSE_QTY = 12
MIN_COVERAGE_DAYS = 60
NEED_DAYS = 30
EXPIRE_LIMIT_DAYS = 120
QUARTER_DAYS = 120

EXCLUDED_KEYWORDS = []

def load_branch(path):
    df = pd.read_excel(path)

    if EXCLUDED_KEYWORDS:
        pattern = "|".join(EXCLUDED_KEYWORDS)
        df = df[~df["ITEM_NAME"].str.contains(pattern, case=False, na=False)]

    df["DAILY_CONSUMPTION"] = df["SALES_QTY"] / QUARTER_DAYS
    df["NEED_30"] = df["DAILY_CONSUMPTION"] * NEED_DAYS
    df["MIN_COVERAGE_QTY"] = df["DAILY_CONSUMPTION"] * MIN_COVERAGE_DAYS
    df["EXCESS"] = df["BALANCE"] - df["MIN_COVERAGE_QTY"]
    df["EXCESS"] = df["EXCESS"].apply(lambda x: x if x > 0 else 0)

    today = datetime.today()
    df["EXPIRE_DATE"] = pd.to_datetime(df["EXPIRE_DATE"])
    df["DAYS_TO_EXPIRE"] = (df["EXPIRE_DATE"] - today).dt.days
    df = df[df["DAYS_TO_EXPIRE"] > EXPIRE_LIMIT_DAYS]

    branch_name = df["STORE_NAME"].iloc[0]
    return branch_name, df


def load_warehouse(path):
    df = pd.read_excel(path)
    df = df[["ITEM_CODE", "ST_BAL"]]
    df.rename(columns={"ST_BAL": "WAREHOUSE_BALANCE"}, inplace=True)
    return df


def generate_transfer(source_name, target_name, source_df, target_df, warehouse_df):

    merged = source_df.merge(target_df, on="ITEM_CODE", suffixes=("_SRC", "_TGT"), how="left")
    merged = merged.merge(warehouse_df, on="ITEM_CODE", how="left")

    merged["WAREHOUSE_BALANCE"].fillna(0, inplace=True)
    merged = merged[merged["WAREHOUSE_BALANCE"] <= MIN_WAREHOUSE_QTY]

    merged["TRANSFER_QTY"] = merged.apply(
        lambda row: min(row["EXCESS_SRC"], row["NEED_30_TGT"])
        if row["EXCESS_SRC"] > 0 and row["NEED_30_TGT"] > 0 else 0,
        axis=1
    )

    merged["TRANSFER_QTY"] = merged["TRANSFER_QTY"].round().astype(int)
    merged = merged[merged["TRANSFER_QTY"] >= 2]

    if merged.empty:
        return None

    def get_reason(row):

        if pd.isna(row["BALANCE_TGT"]):
            return "الصنف غير موجود في الفرع المستهدف"

        if row["BALANCE_TGT"] == 0:
            if row["SALES_QTY_TGT"] > 0:
                return "الصنف مصفّر في الفرع المستهدف – و عليه طلب"
            else:
                return "الصنف مصفّر في الفرع المستهدف – و بدون حركة"

        if row["NEED_30_TGT"] > row["BALANCE_TGT"]:
            return "الصنف عليه طلب عالي في الفرع المستهدف"

        return "الصنف زيادة في الفرع المصدر"

    merged["REASON"] = merged.apply(get_reason, axis=1)

    final_report = merged[[
        "ITEM_CODE",
        "ITEM_NAME_SRC",
        "TRANSFER_QTY",
        "EXPIRE_DATE_SRC",
        "REASON",
        "NEED_30_TGT"
    ]]

    final_report.rename(columns={
        "ITEM_NAME_SRC": "ITEM_NAME",
        "EXPIRE_DATE_SRC": "EXPIRE_DATE"
    }, inplace=True)

    return final_report


def save_report(df, source_name, target_name):
    folder = "Branch_Rebalancing_Results"
    if not os.path.exists(folder):
        os.makedirs(folder)

    file_name = f"تحويل من {source_name} إلى {target_name}.xlsx"
    path = os.path.join(folder, file_name)

    df.to_excel(path, index=False)

    wb = load_workbook(path)
    ws = wb.active

    ref = f"A1:E{ws.max_row}"
    table = Table(displayName="RebalancingTable", ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    yellow = PatternFill(start_color="FFFABF", end_color="FFFABF", fill_type="solid")
    gray = PatternFill(start_color="E7E7E7", end_color="E7E7E7", fill_type="solid")
    red = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
        cell = row[0]
        text = str(cell.value)
        if "طلب عالي" in text:
            cell.fill = yellow
        elif "زيادة" in text:
            cell.fill = gray
        elif "مصفّر" in text:
            cell.fill = red

    wb.save(path)


warehouse_df = load_warehouse(warehouse_path)

branches = []
for path in branch_files:
    name, df = load_branch(path)
    branches.append((name, df))

pairs = list(itertools.combinations(branches, 2))

raw_candidates = {}

for (nameA, dfA), (nameB, dfB) in pairs:

    report_A_to_B = generate_transfer(nameA, nameB, dfA, dfB, warehouse_df)
    if report_A_to_B is not None:
        for _, row in report_A_to_B.iterrows():
            raw_candidates.setdefault(row["ITEM_CODE"], []).append({
                "source": nameA,
                "target": nameB,
                "row": row
            })

    report_B_to_A = generate_transfer(nameB, nameA, dfB, dfA, warehouse_df)
    if report_B_to_A is not None:
        for _, row in report_B_to_A.iterrows():
            raw_candidates.setdefault(row["ITEM_CODE"], []).append({
                "source": nameB,
                "target": nameA,
                "row": row
            })

reports = {}

for item_code, candidates in raw_candidates.items():
    best = max(candidates, key=lambda c: c["row"]["NEED_30_TGT"])
    key = (best["source"], best["target"])
    reports.setdefault(key, []).append(best["row"])

for (source_name, target_name), rows in reports.items():
    df_out = pd.DataFrame(rows)
    df_out = df_out.drop(columns=["NEED_30_TGT"])
    save_report(df_out, source_name, target_name)

print(f"✔ تم إنشاء {len(reports)} ملف تحويل، بإجمالي {len(raw_candidates)} صنف.")
